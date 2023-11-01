# Created By: Cameron Beebe
# Created on: 11/1/23
# Purpose: Aid in creating a mileage spreadsheet when traveling between plants

import tkinter as tk
from tkinter import messagebox
import openpyxl
from datetime import date
import os
import subprocess

# Constants
LOCATIONS = {
    "Warren": {"Plant 1": 23, "Plant 2": 23.7, "Orwell": 23},
    "Plant 1": {"Warren": 23, "Plant 2": 1.2, "Orwell": 17},
    "Plant 2": {"Warren": 23.7, "Plant 1": 1.2, "Orwell": 18.3},
    "Orwell": {"Warren": 23, "Plant 1": 17, "Plant 2": 18.3},
}

# Variables
location_to = ""
distance = 0

def submit():
    global location_to, distance
    selected_option = radio_var.get()
    location_to = selected_option
    
    # Update label text and radio button options
    text_label.config(text="Where are you driving from?")
    radio_var.set("Warren")
    for i, option in enumerate(options):
        radio_buttons[i].config(text=option)
    radio_buttons[-1].pack(anchor=tk.W)
    
    # Update submit button command
    submit_button.config(command=submit_location_from)

def submit_location_from():
    global distance
    selected_option_from = radio_var.get()
    
    if location_to in LOCATIONS and selected_option_from in LOCATIONS[location_to]:
        distance = LOCATIONS[location_to][selected_option_from]
    
    messagebox.showinfo(
        "Submission",
        f"You selected:\nLocation To: {location_to}\nLocation From: {selected_option_from}\nDistance: {distance} miles"
    )
    is_round_trip() # Call the is_round_trip function 
    open_excel_sheet(location_to, selected_option_from, distance)


def open_excel_sheet(location_to, location_from, distance):
    file_name = "Mileage.xlsx"

    # Check if the file already exists
    if os.path.exists(file_name):
        # Open the existing workbook
        wb = openpyxl.load_workbook(file_name)
        # Select the active sheet
        sheet = wb.active
        # Find the last row in the sheet
        last_row = sheet.max_row
        # Add the new data in the next row
        sheet.cell(row=last_row + 1, column=1, value=location_to)
        sheet.cell(row=last_row + 1, column=2, value=location_from)
        sheet.cell(row=last_row + 1, column=3, value=distance)
        sheet.cell(row=last_row + 1, column=4, value=date.today())
    else:
        # Create a new workbook
        wb = openpyxl.Workbook()
        # Select the active sheet
        sheet = wb.active
        # Set the column headers
        sheet['A1'] = 'Location To'
        sheet['B1'] = 'Location From'
        sheet['C1'] = 'Distance (miles)'
        sheet['D1'] = 'Date'
        # Set the data
        sheet['A2'] = location_to
        sheet['B2'] = location_from
        sheet['C2'] = distance
        sheet['D2'] = date.today()

    # Save the workbook
    wb.save(file_name)

    # Open the file using the default program
    subprocess.Popen(['start', file_name], shell=True)

    # Close the program
    window.quit()

def is_round_trip():
    global distance
    answer = messagebox.askquestion("Round Trip", "Is this a round trip?")
    if answer == "yes":
        distance *= 2


# Create the main window
window = tk.Tk()
window.title("Driving Survey")

# Set the window size
window_width = 400
window_height = 175

# Get the screen dimensions
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()

# Calculate the window position
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2

# Set the window position
window.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Create the text label
text_label = tk.Label(window, text="Where did you drive today?")
text_label.pack()

# Create the radio buttons
radio_var = tk.StringVar()
radio_var.set("Warren")

radio_frame = tk.Frame(window)
radio_frame.pack()

options = ["Warren", "Plant 1", "Plant 2", "Orwell"]
radio_buttons = []

# List to store the radio buttons
for option in options:
    radio_button = tk.Radiobutton(radio_frame, text=option, variable=radio_var, value=option)
    radio_button.pack(anchor=tk.W)
    radio_buttons.append(radio_button)

# Create the submit button
submit_button = tk.Button(window, text="Submit", command=submit)
submit_button.pack()

# Create the GUI event loop
window.mainloop()


